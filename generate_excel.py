#!/usr/bin/env python3
"""
Generator profesjonalnego arkusza Excel do oceny kandydatow
Wykonaj: python generate_excel.py
Uzywamy czcionek standardowych dostepnych we wszystkich wersjach Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

def generate_evaluation_spreadsheet():
    """Generuje plik Excel Ocena_Kandydatow.xlsx"""
    
    # Tworzenie nowego skoroszytu
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ocena kandydatow"

    # Definicja naglowkow
    headers = [
        "ID kandydata", "Imie", "Nazwisko", "Stanowisko", "Data rozmowy",
        "Wiedza techniczna", "Microsoft 365", "Rozwiazywanie problemow",
        "Cyberbezpieczenstwo", "Komunikacja", "Wspolpraca", "Motywacja",
        "Wynik", "Rekomendacja", "Uwagi"
    ]

    # Formatowanie naglowkow - uzywamy Arial zamiast Calibri
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Dodanie naglowkow
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Szerokosci kolumn
    column_widths = [10, 12, 12, 15, 12, 12, 12, 14, 14, 12, 12, 10, 10, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(idx)].width = width

    # Zamrozenie naglowka
    ws1.freeze_panes = "A2"

    # Dodanie 30 pustych wierszy z ramkami
    for row in range(2, 32):
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name='Arial', size=10)

    # Walidacja danych dla ocen (kolumny F-L, czyli 6-12)
    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv.error = 'Wpisz 1-5'
    dv.errorTitle = 'Nieprawidlowa wartosc'
    ws1.add_data_validation(dv)

    for row in range(2, 32):
        for col in range(6, 13):
            dv.add(ws1.cell(row=row, column=col))

    # Formuly dla wyniku koncowego (kolumna M, nr 13)
    for row in range(2, 32):
        cell = ws1.cell(row=row, column=13)
        cell.value = '=IF(COUNTA(F' + str(row) + ':L' + str(row) + ')=0,"",ROUND(AVERAGE(F' + str(row) + ':L' + str(row) + '),2))'
        cell.number_format = '0.00'
        cell.font = Font(name='Arial', size=10, bold=True)

    # Formuly dla rekomendacji (kolumna N, nr 14)
    for row in range(2, 32):
        cell = ws1.cell(row=row, column=14)
        cell.value = '=IF(M' + str(row) + '="","",IF(M' + str(row) + '>=4.5,"Zatrudnic",IF(M' + str(row) + '>=3.5,"Rozwazyc","Odrzucic")))'
        cell.font = Font(name='Arial', size=10, bold=True)

    # Formatowanie warunkowe
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    green_font = Font(name='Arial', size=10, color="006100", bold=True)
    yellow_font = Font(name='Arial', size=10, color="9C6500", bold=True)
    red_font = Font(name='Arial', size=10, color="9C0006", bold=True)

    # Reguly formatowania dla wyniku
    ws1.conditional_formatting.add('M2:M31', CellIsRule(operator='greaterThanOrEqual', formula=['4.5'], fill=green_fill, font=green_font))
    ws1.conditional_formatting.add('M2:M31', CellIsRule(operator='greaterThanOrEqual', formula=['3.5'], fill=yellow_fill, font=yellow_font))
    ws1.conditional_formatting.add('M2:M31', CellIsRule(operator='lessThan', formula=['3.5'], fill=red_fill, font=red_font))

    # Reguly formatowania dla rekomendacji
    ws1.conditional_formatting.add('N2:N31', CellIsRule(operator='equal', formula=['"Zatrudnic"'], fill=green_fill, font=green_font))
    ws1.conditional_formatting.add('N2:N31', CellIsRule(operator='equal', formula=['"Rozwazyc"'], fill=yellow_fill, font=yellow_font))
    ws1.conditional_formatting.add('N2:N31', CellIsRule(operator='equal', formula=['"Odrzucic"'], fill=red_fill, font=red_font))

    # Filtry AutoFilter
    ws1.auto_filter.ref = 'A1:' + get_column_letter(len(headers)) + '31'

    # ==================== ARKUSZ 2: PYTANIA ====================
    ws2 = wb.create_sheet("Pytania rekrutacyjne")

    # Naglowek
    ws2.merge_cells('A1:B1')
    title = ws2['A1']
    title.value = "PYTANIA REKRUTACYJNE"
    title.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25

    # Podtytul
    ws2.merge_cells('A2:B2')
    subtitle = ws2['A2']
    subtitle.value = "Kompetencje Techniczne"
    subtitle.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    subtitle.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2.row_dimensions[2].height = 20

    # Pytania
    questions = [
        "1. Jakimi uslugami Microsoft 365 zarzadzal(a)s w poprzednich rolach?",
        "2. Jak zabezpieczylbys srodowisko Microsoft 365?",
        "3. Jakie roznice widzisz miedzy SharePoint, OneDrive i Teams?",
        "4. Jak diagnozujesz problemy z logowaniem uzytkownikow?",
        "5. Opowiedz o najtrudniejszym problemie technicznym, ktorego rozwiazales.",
    ]

    current_row = 3
    for i, question in enumerate(questions, 1):
        ws2.merge_cells('A' + str(current_row) + ':B' + str(current_row))
        cell = ws2['A' + str(current_row)]
        cell.value = question
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border
        ws2.row_dimensions[current_row].height = 35
        current_row += 1

    # Sekcja uwag
    current_row += 1
    ws2.merge_cells('A' + str(current_row) + ':B' + str(current_row))
    notes_header = ws2['A' + str(current_row)]
    notes_header.value = "UWAGI I OBSERWACJE"
    notes_header.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    notes_header.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws2.row_dimensions[current_row].height = 20
    current_row += 1

    for i in range(5):
        ws2.merge_cells('A' + str(current_row) + ':B' + str(current_row))
        cell = ws2['A' + str(current_row)]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.font = Font(name='Arial', size=10)
        ws2.row_dimensions[current_row].height = 30
        current_row += 1

    ws2.column_dimensions['A'].width = 80
    ws2.column_dimensions['B'].width = 2

    # Zapis pliku
    output_file = "Ocena_Kandydatow.xlsx"
    wb.save(output_file)
    print("Plik Excel wygenerowany: " + output_file)
    print("Rozmiar: " + str(os.path.getsize(output_file)) + " bajtow")
    print("Arkusze: Ocena kandydatow i Pytania rekrutacyjne")

if __name__ == "__main__":
    generate_evaluation_spreadsheet()
